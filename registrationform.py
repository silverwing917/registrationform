#Python program to save registration form info into an excel file

from openpyxl import load_workbook
from tkinter import *

# Load Excel workbook
wb = load_workbook('C:\\Users\\sharo\\OneDrive\\Documents\\Desktop\\HIM Stuff\\Resume2025\\Portofolio\\Python\\reg.xlsx')
ws = wb.active

# Set header row
def init_excel():
    headers = ["Name", "Course", "Semester", "Form No.", "Contact Number", "Email", "Address"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i).value = h
    wb.save('C:\\Users\\sharo\\OneDrive\\Documents\\Desktop\\HIM Stuff\\Resume2025\\Portofolio\\Python\\reg.xlsx')

# Clear all fields
def clear_fields():
    for entry in entries:
        entry.delete(0, END)

# Add form data
def insert_data():
    if all(f.get() for f in entries):
        row = ws.max_row + 1
        for i, f in enumerate(entries, 1):
            ws.cell(row=row, column=i).value = f.get()
        wb.save('C:\\Users\\sharo\\OneDrive\\Documents\\Desktop\\HIM Stuff\\Resume2025\\Portofolio\\Python\\reg.xlsx')
        clear_fields()
        print("Data saved successfully.")
    else:
        print("Please fill all fields.")

# GUI setup
root = Tk()
root.title("Registration Form")

labels = ["Name", "Course", "Semester", "Form No.", "Contact Number", "Email", "Address"]
entries = []

for i, label_text in enumerate(labels):
    label = Label(root, text=label_text)
    label.grid(row=i, column=0, padx=10, pady=5, sticky=E)
    entry = Entry(root, width=40)
    entry.grid(row=i, column=1, padx=10, pady=5)
    entries.append(entry)

submit_btn = Button(root, text="Submit", command=insert_data)
submit_btn.grid(row=len(labels), column=0, columnspan=2, pady=10)

init_excel()
root.mainloop()



